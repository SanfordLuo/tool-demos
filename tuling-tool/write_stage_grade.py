"""
写入单词适用的年级
"""

import json
from openpyxl import load_workbook

v3_path = "D:\\CodeOkay\\tuling-tool\\data\\xlsx\\neworiental_v3.xlsx"
json_path = "D:\\CodeOkay\\tuling-tool\\utils\\word_stage_grade.json"


def load_data():
    sheet_name = "Tablib Dataset"
    begin_idx = 2
    end_idx = 463736

    ret_dict = {}

    wb = load_workbook(v3_path)
    ws = wb[sheet_name]
    rows = ws[begin_idx:end_idx]

    tag = {1: "初中", 2: "高中", 3: "小学"}

    for row in rows:
        word = row[0].value
        stage_id = int(row[1].value)
        grade_name = row[5].value
        if word in ret_dict:
            if tag[stage_id] not in ret_dict[word]["stage_id"]:
                ret_dict[word]["stage_id"].append(tag[stage_id])
            if grade_name not in ret_dict[word]["grade_name"]:
                ret_dict[word]["grade_name"].append(grade_name)
        else:
            ret_dict[word] = {"stage_id": [tag[stage_id]], "grade_name": [grade_name]}

    return ret_dict


def write_json():
    with open(json_path, "r", encoding='utf-8') as f:
        load_dict = json.load(f)

    load_dict.update(ret_dict)
    with open(json_path, "w+", encoding='utf-8') as ff:
        json.dump(load_dict, ff, indent=4, ensure_ascii=False)


if __name__ == '__main__':
    ret_dict = load_data()
    if ret_dict:
        write_json()

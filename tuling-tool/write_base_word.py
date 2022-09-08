import json
from openpyxl import load_workbook


v3_path = "D:\\CodeOkay\\tuling-tool\\data\\xlsx\\neworiental_v3_base_word.xlsx"
ret_path = "D:\\CodeOkay\\tuling-tool\\data\\result\\造课信息汇总0714.xlsx"


def load_v3_data():
    sheet_name = "Result 1"
    begin_idx = 1
    end_idx = 8990

    wb = load_workbook(v3_path)
    ws = wb[sheet_name]
    rows = ws[begin_idx:end_idx]
    base_words = [row[0].value for row in rows]
    return base_words


def write_data():
    sheet_name = "单词"
    begin_idx = 2
    end_idx = 7611
    write_column = 6

    wb = load_workbook(ret_path)
    ws = wb[sheet_name]
    rows = ws[begin_idx:end_idx]
    for idx, row in enumerate(rows):
        if row[1].value in base_words:
            ws.cell(row=begin_idx+idx, column=write_column, value=1)
        else:
            ws.cell(row=begin_idx+idx, column=write_column, value=0)

    wb.save(ret_path)


if __name__ == '__main__':
    base_words = load_v3_data()
    write_data()

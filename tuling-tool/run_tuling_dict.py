import os
import logging
import asyncio
from utils.util_online_dict import OnlineDictHandler
from openpyxl import load_workbook
from utils import util_logger
from utils import const

loger_name = os.path.basename(__file__).split(".")[0]
util_logger.set_logger_config(loger_name)
logger = logging.getLogger(loger_name)

xlsx_path = "D:\\CodeOkay\\tuling-tool\\data\\xlsx\\test英语单词.xlsx"
txt_path = "D:\\CodeOkay\\tuling-tool\\data\\test英文单词_1_5000.txt"
sheet_name = "英文单词"
begin_idx = 1
end_idx = 1000

dict_type = const.TULING_DICT
dict_dicts = [const.DICT_EN_WORD]


def load_data():
    old_list = []
    if os.path.exists(txt_path):
        with open(txt_path, encoding='utf-8') as f1:
            txt_data = f1.readlines()
        for _txt in txt_data:
            old_list.append(int(_txt.split(",", 1)[0]))

    data = {}
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]
    rows = ws[begin_idx:end_idx]
    for _k, _v in enumerate(rows):
        if _v[0] and (_k+begin_idx) not in old_list:
            data[_k+begin_idx] = _v[0].value

    wb.close()

    logger.info(f"load_data nums: {len(data)}")
    return data


async def write_data(w_data):
    w_file_path = f"D:\\CodeOkay\\tuling-tool\\data\\{sheet_name}_{begin_idx}_{end_idx}.txt"
    with open(w_file_path, "a", encoding="utf-8") as f:
        f.write(w_data + "\n")


async def run_translate():
    req_data = load_data()
    online_dict_handler = OnlineDictHandler(dict_type, dict_dicts)

    while req_data:
        del_key = []
        for k, text in req_data.items():
            ret_text = await online_dict_handler.translate(text)
            if ret_text:
                del_key.append(k)
                await write_data(f"{k},{text},{ret_text}")

        for _k in del_key:
            if _k in list(req_data.keys()):
                del req_data[_k]


if __name__ == '__main__':
    logger.info(f"[begin] === sheet_name:{sheet_name}, begin_idx:{begin_idx}, end_idx:{end_idx}")
    asyncio.run(run_translate())
    logger.info(f"[end] === sheet_name:{sheet_name}, begin_idx:{begin_idx}, end_idx:{end_idx}")

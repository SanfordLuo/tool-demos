from openpyxl import load_workbook

# xlsx_file = None
xlsx_file = "D:\\CodeOkay\\tuling-tool\\data\\xlsx\\造课key跑数据_结果初筛7.14 (1).xlsx"
txt_file = "D:\\CodeOkay\\tuling-tool\\data\\result\\造课key跑数据_结果初筛7.14 (1)_单词_2_6868.txt"
ret_xlsx_file = "D:\\CodeOkay\\tuling-tool\\data\\result\\造课key跑数据_结果初筛7.14 (1).xlsx"

sheet_name = "单词"
write_column = 7


def write_data():
    with open(txt_file, encoding='utf-8') as txt_f:
        txt_data = txt_f.readlines()

    read_list = []
    for _txt in txt_data:
        if _txt:
            read_list.append(_txt.split(",", 2))

    if xlsx_file:
        wb = load_workbook(xlsx_file)
    else:
        wb = load_workbook(ret_xlsx_file)

    ws = wb[sheet_name]

    print(f"read_list: {len(read_list)}")
    for _data in read_list:
        ws.cell(row=int(_data[0]), column=write_column, value=_data[2])

    wb.save(ret_xlsx_file)


if __name__ == '__main__':
    print("===== begin =====")
    write_data()
    print("===== end =====")

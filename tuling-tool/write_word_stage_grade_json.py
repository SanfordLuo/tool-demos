import json
from openpyxl import load_workbook


json_path = "D:\\CodeOkay\\tuling-tool\\utils\\word_stage_grade.json"
ret_path  = "D:\\CodeOkay\\tuling-tool\\data\\xlsx\\造课信息汇总0714.xlsx"

sheet_name = "单词"
word_column = 2
stage_id_column = 4
grade_name_column = 5
begin_idx = 2
end_idx = 7611


def load_data():
    with open(json_path, "r", encoding='utf-8') as f:
        load_dict = json.load(f)
    return load_dict


def write_ret():
    wb = load_workbook(ret_path)
    ws = wb[sheet_name]
    rows = ws[begin_idx:end_idx]
    print(f"===== {len(rows)} =====")
    for idx, row in enumerate(rows):
        print(f"===== {idx} =====")
        word = row[word_column-1].value
        if word in ret_dict:
            stage_id = ", ".join(ret_dict[word]["stage_id"])
            grade_name = ", ".join(ret_dict[word]["grade_name"])
            ws.cell(row=begin_idx+idx, column=stage_id_column, value=stage_id)
            ws.cell(row=begin_idx+idx, column=grade_name_column, value=grade_name)

    wb.save(ret_path)


if __name__ == '__main__':
    ret_dict = load_data()
    write_ret()

import os
import re
import json
import logging
from openpyxl import load_workbook
from utils import util_logger

loger_name = os.path.basename(__file__).split(".")[0]
util_logger.set_logger_config(loger_name)
logger = logging.getLogger(loger_name)

file_path = "D:\\CodeOkay\\tuling-tool\\data\\xlsx\\英语单词.xlsx"
sheet_name = "英文单词"
col = 2

begin_row = 1
end_row = 1000


def load_data():
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    rows = ws[begin_row:end_row]
    for idx, row in enumerate(rows):
        if row[col-1]:
            row_num = idx+begin_row
            val = row[col-1].value
            try:
                insert_db(row_num, val)
            except Exception as e:
                logger.error(f"写入数据失败. row_num: {row_num}, {e}")


def insert_db(row_num, val):
    logger.info(f"开始写入数据 row_num: {row_num}")

    val_dict = json.loads(val)
    info = val_dict.get("nlpResponse", {}).get("intent", {}).get("parameters", {}).get("result", {}).get("info", {})
    if info:
        explanations = info.get("explanations", [])
        explanations_dict = {}
        for explanation in explanations:
            if explanation["pos"] in explanations_dict:
                explanations_dict[explanation["pos"]] += f"；{explanation['meaning']}"
            else:
                explanations_dict[explanation["pos"]] = explanation['meaning']

        word = {
            "word": info.get("word", ""),
            "phonetic_en": info.get("phoneticEn", ""),
            "phonetic_us": info.get("phoneticUs", ""),
            "explanations": json.dumps(explanations_dict),
            "word_audio": info.get("wordAudio", ""),
            "phonetic_en_audio": info.get("phoneticEnAudio", ""),
            "phonetic_us_audio": info.get("phoneticUsAudio", ""),
            "noun_plurals": info.get("nounPlurals", ""),
            "verb": info.get("verb", ""),
            "third_singular": info.get("thirdSingular", ""),
            "preterit": info.get("preterit", ""),
            "present_participle": info.get("presentParticiple", ""),
            "past_participle": info.get("pastParticiple", ""),
            "adjective": info.get("adjective", ""),
            "comparative_degree": info.get("comparativeDegree", ""),
            "superlative_degree": info.get("superlativeDegree", ""),
            "word_status": 1
        }
        word_args = list(word.values())
        word_id = 000000

        examples = info.get("examples", [])
        example_args = []
        for example in examples:
            example_args.append([word_id, example["en"], example["cn"]])

    logger.info(f"结束写入数据 row_num: {row_num}")


SQL_INSERT_EN_WORD = """

"""

SQL_INSERT_EN_WORD_EXAMPLE = """

"""
